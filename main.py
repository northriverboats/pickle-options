#!/usr/bin/env python

from PyQt4 import QtGui # Import the PyQt4 module we'll need
from PyQt4.QtCore import QSettings, QSize, QPoint, QCoreApplication, QThread, SIGNAL # pylint: disable-msg=E0611
from pathlib import Path
from dotenv import load_dotenv
from fields import start, end, width, sections
from fields import topSection, costSummary, bottomSection
from fields import startSections, startSectionsSize, endSections, partSection, partSectionByModel
import openpyxl
import pickle
import sys # We need sys so that we can pass argv to QApplication
import tempfile
import os
import re
import click
import MainWindow  # This file holds our MainWindow and all design related things



r"""
Notes:

To design UI: Lib\site-packages\PyQt4\Designer.exe

To rebuild UI: Lib\site-packages\PyQt4\pyuic4 MainWindow.ui  -o MainWindow.py

Developed in C:\\Development\\nrb-pickle-boats :
venv\Scripts\pyinstaller.exe --onefile --windowed --icon options.ico  --name "Options Pickler" "Pickle Options FWW.spec" main.py

ToDo's
"""

class MainAppWindow(QtGui.QMainWindow, MainWindow.Ui_MainWindow):

    def __init__(self):
        super().__init__()
        self.setupUi(self)

        # set python environment
        if getattr(sys, 'frozen', False):
            bundle_dir = sys._MEIPASS # pylint: disable=no-member
        else:
            # we are running in a normal Python environment
            bundle_dir = os.path.dirname(os.path.abspath(__file__))

        # load environmental variables
        load_dotenv(dotenv_path = Path(bundle_dir) / ".env")

        # set program icon
        self.setWindowIcon(QtGui.QIcon(os.path.join(bundle_dir, "pickle.ico")))

        # work in INI File Stuff here
        QCoreApplication.setOrganizationName("NRB")
        QCoreApplication.setOrganizationDomain("northriverboats.com")
        QCoreApplication.setApplicationName("Options Boat Pickler")
        self.settings = QSettings()
        

        # set variables
        self.background_thread = None
        self.exit_flag = False
        self.dir = self.settings.value("dir", os.getenv("DIR"))
        self.pickle_name = os.getenv("PICKLE")

        # set ui state
        self.actionCancel.setEnabled(False)
        self.btnCancel.hide()
        self.lePath.setText(self.dir)

        # set slots and signals
        self.actionExit.triggered.connect(self.closeEvent)
        self.actionAbout.triggered.connect(self.doAbout)
        self.btnBrowse.clicked.connect(self.browseEvent)
        self.btnRun.clicked.connect(self.startBackgroundTask)

    def doAbout(self, event):
        about_msg = "NRB Boat Folder Pickler\n©2019 North River Boats\nBy Fred Warren"
        QtGui.QMessageBox.information(self, 'About',
                about_msg, QtGui.QMessageBox.Ok)

    def closeEvent(self, e):
        self._closeEvent(0)

    def _closeEvent(self, e):
        self.exit_flag = True
        self.settings.setValue("dir", self.dir)
        sys.exit(0)

    def browseEvent(self):
        default_dir = self.dir
        my_dir = QtGui.QFileDialog.getExistingDirectory(self, "Open a folder", default_dir, QtGui.QFileDialog.ShowDirsOnly)
        if my_dir != "":
            self.dir = my_dir
            self.update_progressbar(0)
        self.lePath.setText(self.dir)



    def startBackgroundTask(self):
        # hide / disable buttons and menu items as well as saving state
        self.block_actions()
        self.background_thread = background_thread(self.dir)

        self.connect(self.background_thread, SIGNAL('endBackgroundTask()'), self.endBackgroundTask)
        self.connect(self.background_thread, SIGNAL('update_statusbar(QString)'), self.update_statusbar)
        self.connect(self.background_thread, SIGNAL('update_label(QString)'), self.update_label)
        self.connect(self.background_thread, SIGNAL('update_progressbar(int)'), self.update_progressbar)
        self.btnCancel.clicked.connect(self.doAbort)    
        # Thread will self-terminate or be stopped via update_abort
        self.background_thread.start()

    def endBackgroundTask(self):
        self.unblock_actions()
        self.statusbar.showMessage("")

    def doAbort(self):
        self.background_thread.running = False
        self.unblock_actions

    def block_actions(self):        
        self.btnRun.setEnabled(False)
        self.btnCancel.setEnabled(True)
        self.btnRun.hide()
        self.btnCancel.show()
        self.actionRun.setEnabled(False)
        self.actionCancel.setEnabled(True)
        self.lePath.setReadOnly(True)
    
    def unblock_actions(self):
        self.btnRun.setEnabled(True)
        self.btnCancel.setEnabled(False)
        self.btnRun.show()
        self.btnCancel.hide()
        self.actionRun.setEnabled(True)
        self.actionCancel.setEnabled(False)
        self.lePath.setReadOnly(False)
        self.lblFile.setText("")

    def update_statusbar(self, message):
        self.statusbar.showMessage(message)

    def update_progressbar(self, num):
        self.progressBar.setValue(num)
        
    def update_label(self, label):
        self.lblFile.setText(label)


class pickler():
    def __init__(self, dir):
        self.dir = dir
        self.wb = None
        self.ws = None
        self.starts = []
        self.ends = []
        self.boatSizes = []
        self.data = {}
        self.part = {}
        self.parts = {}
        self.not_found = []
        self.dupes = []

    def build_files_list(self):
        files = []
        for path in Path(self.dir).glob("[!~$]*.xlsx"):
            files.append(path)
        return files

    def open_sheet(self, file):
        self.wb = openpyxl.load_workbook(file, data_only = True)
        self.ws = self.wb.active

    def output_not_found(self):
        output = ""
        for error in sorted(self.not_found, key=lambda x: [x[0], x[1], x[2]]):
            output += " {:30.30}    {:12.12}    {:30.30}\n".format(error[0], error[1], error[2])
        return output

    def output_dupes(self):
        output = ""
        for error in sorted(self.dupes, key=lambda x: [x[0], x[1], x[2]]):
            output += " {:30.30}    {:12.12}    {:30.30}\n".format(str(error[0]), str(error[1]), str(error[2]))
        return output


    #### IDENTIFY PORTIONS OF SHEET TO PROCESS #############
    def find_starts(self):
        # find where sections start
        self.starts = []
        for row in self.ws.iter_cols(min_col=start, max_col=start):
            for cell in row:
                if cell.value == "QTY":
                    self.starts.append(cell.row)

    def find_ends(self):
        # find where sections end
        self.ends = []
        for row in self.ws.iter_cols(min_col=end, max_col=end):
            for cell in row:
                if cell.value == "CREDIT - TOTAL":
                    self.ends.append(cell.row)
    
    def find_boat_sizes(self):
        self.boatSizes = []
        for col in self.ws.iter_rows(min_row=1, max_row=1):
            for cell in col:
                if  cell.column > 3 and cell.data_type == 'n' and cell.value != None:
                    self.boatSizes.append(cell.value)
	
    def find_picklest(self):
        global partSection
        picklist = self.boatSizes[-1] + width
        partSection[-1][1] = picklist
    
    #### PROCESS NON-SECTION PORTIONS OF THE SHEET #############    
    def process_top_section(self):
        # Process top static section of sheet
        for name, column, row, default in topSection:
            value = self.ws.cell(column = column, row = row).value
            if value is None:
                value = default
            self.data[name] = value

    def process_cost_summary_by_boat_size(self):
        #Process cost summary
        for i, boatSize in enumerate(self.boatSizes):
            for name, column, row, default in costSummary:
                value = self.ws.cell(column = column + (i * width), row = row).value
                if value is None:
                    value = default
                self.data[str(boatSize) + name] = value
				
    def process_bottom_section(self):
        # Process bottom section
        offset = self.ends[4] + 6
        for name, column, row, default in bottomSection:
            value = self.ws.cell(column = column, row = row + offset).value
            if value is None:
                value = default
            self.data[name] = value				

    #### PROCESS SECTION PORTIONS OF THE SHEET SUPPORTING FUNCTIONS #############
    def process_inner_section_top(self, index, offset, section):
        for name, column, row, default in startSections:
            value = self.ws.cell(column = column, row = row + offset).value
            if value is None:
                value = default
            self.data[section + name] = value

    def process_inner_section_top_by_boat_size(self, index, offset, boatSize, section):
        for name, column, row, default in startSectionsSize:
            value = self.ws.cell(column = column + (index * width), row = row + offset).value
            if value is None:
                value = default
            self.data[section + " " + str(boatSize) + name] = value

    def process_inner_section_bottom_by_boat_size(self, index, offset, boatSize, section):
        for name, column, row, default in endSections:
            value = self.ws.cell(column = column + (index * width), row = row + offset).value
            if value is None:
                value = default
            self.data[section + " " + str(boatSize) + name] = value


    def process_part_by_boat_size(self, index, offset, boatSize, section):
        for name, column, row, default in partSectionByModel:
            value = self.ws.cell(column = column + (index * width), row = row + offset).value
            if value is None:
                value = default
            self.part[str(boatSize) + name] = value

    def process_part_by_non_boat_size(self, offset, section):
        for name, column, row, default in partSection:
            value = self.ws.cell(column = column, row = row + offset).value
            if value is None:
                value = default
            self.part[name] = value


    def process_section_by_boat_sizes(self, offset, section, process_by_size_function):
        for index, boatSize in enumerate(self.boatSizes):
            process_by_size_function(index, offset, boatSize, section)

    def process_part(self, index, section, offset):
        self.part = {}
        part = self.ws.cell(column = 1, row = 1 + offset).value
        description = self.ws.cell(column = 2, row = 1 + offset).value
        if description == "NOT FOUND":
            error = [
                self.data['FILE'],
                section,
                part,
            ]
            self.not_found.append(error)
        elif part is not None:
            self.parts[part] = self.parts.get(part, 0) + 1
            self.process_part_by_non_boat_size(offset, section)
            self.process_section_by_boat_sizes(offset, section, self.process_part_by_boat_size)
            self.data[section + " PARTS"].append(self.part)

    def process_parts(self, index, section):
        self.parts = {}
        for offset in range(self.starts[index], self.ends[index]-1):
            self.process_part(index, section, offset)
        for part in {part for part in self.parts if self.parts[part] > 1}:
            error = [
                self.data['FILE'],
                section,
                part,
            ]
            self.dupes.append(error)
            


    #### PROCESS SECTION PORTIONS OF THE SHEET #############
    def process_section_top(self):
        # Process top non-parts portion of sections not by boat size
        for index, section in enumerate(sections):
            offset = self.starts[index]
            self.process_inner_section_top(index, offset, section)

    def process_section_top_by_boat_size(self):
		# Process top non-parts portion of sections by boat size
        for i, section in enumerate(sections):
            offset = self.starts[i]
            self.process_section_by_boat_sizes(offset, section, self.process_inner_section_top_by_boat_size)
    
    def process_section_parts(self):
        # Process parts portion of sections both by non boat size and by boat size must be combined
        for index, section in enumerate(sections):
            self.data[section + " PARTS"] = []
            self.process_parts(index, section)
    
    def process_section_bottom(self):
        # Process bottom non-parts portion of sections
        for i, section in enumerate(sections):
            offset = self.ends[i]
            self.process_section_by_boat_sizes(offset, section, self.process_inner_section_bottom_by_boat_size)
			
			
			
			


    #### PROCESS FULL SHEET ##########################
    def process_sheet(self, file):
        self.open_sheet(file)

        # find different sections on the sheet
        self.find_starts()
        self.find_ends()
        self.find_boat_sizes()
        self.find_picklest()

        self.data = {}

        # keys not derived from fields.py
        self.data["FILE"] = file.name
        self.data["FULLPATH"] = file.resolve()
        self.data['BOAT SIZES'] = self.boatSizes

        # process top non-section band
        self.process_top_section() # only non-boat-size at top of sheet
        self.process_cost_summary_by_boat_size() # only by-boat-size at top of sheet
        self.process_bottom_section() # only non-boat-size at bottom of sheet
        
        # process section bands
        self.process_section_top() # only non-boat-size
        self.process_section_top_by_boat_size() # only by-boat-size
        self.process_section_parts() # both non-boat-size and by-boat-size
        self.process_section_bottom() # only non-boat-size

        return [str(self.data["OPTION NUMBER"]), self.data, self.output_not_found(), self.output_dupes()]


class background_thread(QThread):
    def __init__(self, dir):
        QThread.__init__(self)
        self.dir = dir

    def __del__(self):
        self.wait()

    def handle_errors(self, not_founds, dupes):
        if len(not_founds) + len(dupes):
            with tempfile.TemporaryFile(suffix='.txt', mode='w', delete=False) as file:
                if len(not_founds):
                    file.write("--- PARTS NOT FOUND --------------------------------------------------------------\n\n")
                    file.write(not_founds)
                    file.write("\n\n")
                if len(dupes):
                    file.write("--- DUPLICATE PARTS --------------------------------------------------------------\n\n")
                    file.write(dupes)
                    file.write("\n")
                os.startfile(file.name)

    def run(self):
        self.running = True
        self.emit(SIGNAL('update_progressbar(int)'), 0)

        pickle_folder = pickler(self.dir)
        self.emit(SIGNAL('update_statusbar(qString)'), "Finding files...")

        files = pickle_folder.build_files_list()
        self.emit(SIGNAL('update_statusbar(qString)'), "Found {} files to process".format(len(files)))

        options = {}
        total_files = len(files)
        current_count = 0
        not_founds = []
        dupes = []

        if not self.running:
            self.emit(SIGNAL('endBackgroundTask()'))
            return

        for file in files:
            if not self.running:
                break

            option, data, not_founds, dupes = pickle_folder.process_sheet(file)
            options[option] = data

            current_count += 1
            self.emit(SIGNAL('update_progressbar(int)'), int(float(current_count) / total_files * 100))
            self.emit(SIGNAL('update_label(QString)'), str(file))
            self.emit(SIGNAL('update_statusbar(QString)'), 'Pickeling %d of %d' % (current_count, total_files))

        # if we get to this point, pickle the results....
        file_name = os.path.join(self.dir, os.path.split(self.dir)[1].lower() + ".pickle")
        pickle.dump(options, open(file_name, 'wb'))
        # handle all_not_found and all_dups
        self.handle_errors(not_founds, dupes)
        self.emit(SIGNAL('endBackgroundTask()'))

def gui():
    app = QtGui.QApplication(sys.argv)  # A new instance of QApplication
    form = MainAppWindow()              # We set the form to be our Main App Wehdiw (design)
    form.show()                         # Show the form
    app.exec_()                         # and execute the app


def cli(folder):
    pickle_folder = pickler(folder)
    click.echo("Finding files...")
    files = pickle_folder.build_files_list()
    click.echo("Found {} files to process".format(len(files)))

    options = {}
    not_founds = []
    dupes = []

    with click.progressbar(files) as bar:
        for file in bar:
            option, data, not_founds, dupes = pickle_folder.process_sheet(file)
            options[option] = data

    file_name = os.path.join(folder, os.path.split(folder)[1].lower() + ".pickle")
    pickle.dump(options, open(file_name, 'wb'))
    click.echo("Saving pickle: {}".format(file_name))
    if len(not_founds):
        click.echo("--- PARTS NOT FOUND --------------------------------------------------------------\n")
        click.echo(not_founds)
        click.echo("")
    if len(dupes):
        click.echo("--- DUPLICATE PARTS --------------------------------------------------------------\n")
        click.echo(dupes)
        click.echo("")


@click.command()
@click.option('--folder', '-f', type=click.Path(exists=True, file_okay=False), help="folder to process")
def main(folder):
    if folder == None:
        gui()
    else:
        cli(folder)

if __name__ == '__main__':              # if we're running file directly and not importing it
    main()                              # pylint: disable=no-value-for-parameter